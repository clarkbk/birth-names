import csv
import os
import re
import xlrd

from database import db, DbBirthRecord, DbYear, MyBirthRecord
from logger import logger
from openpyxl import load_workbook
from peewee import DoesNotExist
from pytictoc import TicToc
from queue import Queue
from threading import Thread


class Processor(object):
    @property
    def data_directory(self):
        return os.path.join('data', self.country)

    def get_datafiles(self):
        datafiles = os.listdir(self.data_directory)
        datafiles.sort(reverse=True)
        datafiles = filter(self.utils.is_valid_file, datafiles)
        return datafiles

    def process(self, q, sentinel):
        t = TicToc()
        files = self.get_datafiles()
        for file in files:

            # skip already-processed files, according to year
            if (int(re.search(r'(\d{4})', file).group(1)) <= 2017):
                continue

            t.tic()
            logger.info(f'{self.country.upper()} file - {file.ljust(12)} - Starting...')

            for record in self.process_file(file):
                br = MyBirthRecord(*record)
                q.put(br.make_db_record())

            time = t.tocvalue()
            logger.info(f'{self.country.upper()} file - {file.ljust(12)} - Finished '
                        f'({time:.1f} seconds)')

        q.put(sentinel)


class USProcessor(Processor):
    country = 'us'

    def process_file(self, file):
        year = int(re.search(r'(\d{4})', file).group(1))

        with open(os.path.join(self.data_directory, file), 'r', encoding='ascii') as f:
            reader = csv.reader(f)

            for row in reader:
                name, sex, births = list(row)

                yield (self.country, year, name, sex, int(births))

    class utils(object):
        @staticmethod
        def is_valid_file(file):
            return file.endswith('.txt')


class UKProcessor(Processor):
    country = 'uk'

    def process_file(self, file):
        year, sex, ext = self.utils.extract_data_from_filename(file)

        if ext == 'xlsx':
            wb = load_workbook(filename=os.path.join(self.data_directory, file))
            for row in wb['Table 6'].iter_rows(min_row=7, max_col=3, values_only=True):
                if not row[0]:
                    break

                name, births = row[1:]

                if births < 5:
                    continue

                yield (self.country, int(year), str(name).title(), sex, births)

        elif ext == 'xls':
            with xlrd.open_workbook(os.path.join(self.data_directory, file)) as book:
                sheet_idx = self.utils.find_correct_worksheet(book)
                sheet = book.sheet_by_index(sheet_idx)
                start_row = self.utils.find_start_row(sheet)

                for i in range(sheet.nrows):
                    if i <= start_row:
                        continue

                    row = [c for c in sheet.row(i) if not self.utils.cell_is_empty(c)]
                    if not self.utils.is_valid_row(row):
                        continue

                    name, births = row[1:]
                    births = int(births.value)

                    if births < 5:
                        continue

                    yield (self.country, int(year), name.value.title(), sex, births)

    class utils(object):
        @staticmethod
        def is_valid_file(file):
            return (
                bool(re.search(r'\.xlsx?$', file)) and
                not file.startswith('historicname')
            )

        @staticmethod
        def extract_data_from_filename(file):
            pttrn = r'\.(xlsx?)$'
            extension = re.search(pttrn, file).group(1)
            file = re.sub(pttrn, '', file)

            data = file.split('_')
            data.append(extension)
            return data

        @staticmethod
        def find_correct_worksheet(book):
            if book.nsheets >= 6:
                pttrn = re.compile('table 6', re.IGNORECASE)
            else:
                pttrn = re.compile('table 3', re.IGNORECASE)
            return [i for i, sht in enumerate(book.sheet_names()) if pttrn.search(sht)][0]

        @staticmethod
        def find_start_row(sheet):
            for i in range(10):
                if all(x in sheet.row_values(i) for x in ['Rank', 'Name']):
                    return i

        @staticmethod
        def is_valid_row(row):
            return (
                len(row) == 3 and
                row[0].ctype == 2 and
                row[1].ctype == 1 and
                row[2].ctype == 2
            )

        @staticmethod
        def cell_is_empty(cell):
            return cell.ctype == 0


class Consumer(object):
    @staticmethod
    def consume(q, sentinel):
        batch = []
        while True:
            while len(batch) < 1000:
                record = q.get()

                if record is sentinel:
                    break

                batch.append(record)
                q.task_done()

            with db.atomic():
                DbBirthRecord.bulk_create(batch)
            batch = []


if __name__ == '__main__':
    logger.info('Starting...')
    q = Queue()
    sentinel = object()

    threads = []
    for processor in [USProcessor, UKProcessor]:
        p = processor()
        t = Thread(name=f'{type(p).__name__}',
                   target=p.process,
                   args=(q, sentinel,),
                   daemon=True)
        t.start()
        threads.append(t)

    t = Thread(name='Consumer',
               target=Consumer.consume,
               args=(q, sentinel,),
               daemon=True)
    t.start()

    for t in threads:
        t.join()

    logger.info('Adding total births per year...')

    for country in ['us', 'uk']:
        with open(os.path.join('data', country, f'births_by_year.csv'), 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                if not row[0].isdigit():
                    continue

                try:
                    y = DbYear.get_by_id(int(row[0]))
                except DoesNotExist:
                    continue

                setattr(y, f'births_{country}_m', row[1])
                setattr(y, f'births_{country}_f', row[2])
                y.save()

    logger.info('Done')
